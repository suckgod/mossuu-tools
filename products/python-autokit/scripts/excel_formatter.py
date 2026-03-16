#!/usr/bin/env python3
"""
Excel Formatter - Standardize Excel tables

Clean and standardize Excel spreadsheets:
- Standardize column names (lowercase, remove spaces, replace special chars)
- Reorder columns (auto-detect or custom order)
- Convert data types (strings to numbers/dates)
- Remove empty rows/columns
- Apply consistent formatting (date, currency, number formats)
- Auto-fit column widths
- Add header styling
- Export to CSV or cleaned Excel

Usage:
    python excel_formatter.py report.xlsx --standardize-headers
    python excel_formatter.py data.xlsx --reorder "Name,Date,Amount,Category"
    python excel_formatter.py sales.xlsx --convert-date "Order Date" --output clean.xlsx
"""

import os
import sys
import re
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import pandas as pd

__version__ = "1.0.0"

class ExcelFormatter:
    def __init__(self, input_file: str,
                 output_file: str = None,
                 standardize_headers: bool = False,
                 reorder_columns: List[str] = None,
                 remove_empty: bool = False,
                 convert_dates: List[str] = None,
                 number_format: str = None,
                 date_format: str = None,
                 auto_width: bool = False,
                 header_style: bool = False,
                 verbose: bool = False):
        self.input_file = Path(input_file)
        self.output_file = Path(output_file) if output_file else self.input_file.parent / f"{self.input_file.stem}_formatted{self.input_file.suffix}"
        self.standardize_headers = standardize_headers
        self.reorder_columns = reorder_columns or []
        self.remove_empty = remove_empty
        self.convert_dates = convert_dates or []
        self.number_format = number_format
        self.date_format = date_format
        self.auto_width = auto_width
        self.header_style = header_style
        self.verbose = verbose

        self.df = None
        self.stats = {
            'original_shape': None,
            'final_shape': None,
            'columns_renamed': 0,
            'rows_removed': 0,
            'columns_removed': 0,
            'dates_converted': 0
        }

    def log(self, msg):
        if self.verbose:
            print(msg)

    def load_excel(self):
        """Load Excel file into DataFrame"""
        self.log(f"Loading: {self.input_file}")
        try:
            self.df = pd.read_excel(self.input_file, engine='openpyxl')
            self.stats['original_shape'] = self.df.shape
            print(f"  Loaded: {self.df.shape[0]} rows × {self.df.shape[1]} columns")
        except Exception as e:
            raise RuntimeError(f"Failed to load Excel: {e}")

    def standardize_column_names(self):
        """Clean and standardize column names"""
        if not self.standardize_headers:
            return

        original_columns = self.df.columns.tolist()
        new_columns = []

        for col in original_columns:
            # Convert to string, lowercase, strip whitespace
            new_col = str(col).strip().lower()

            # Replace spaces and special chars with underscores
            new_col = re.sub(r'[\s\-]+', '_', new_col)  # spaces/hyphens → _
            new_col = re.sub(r'[^\w_]', '', new_col)    # remove non-alphanumeric (except _)

            # Ensure non-empty
            if not new_col:
                new_col = f"column_{len(new_columns)+1}"

            new_columns.append(new_col)

        # Check for duplicates
        if len(set(new_columns)) != len(new_columns):
            # Append numbers to duplicates
            counts = Counter(new_columns)
            final_columns = []
            for col in new_columns:
                if counts[col] > 1:
                    idx = final_columns.count(col) + 1
                    final_columns.append(f"{col}_{idx}")
                else:
                    final_columns.append(col)
            new_columns = final_columns

        self.df.columns = new_columns
        renamed = sum(1 for o, n in zip(original_columns, new_columns) if o != n)
        self.stats['columns_renamed'] = renamed
        print(f"  Standardized {renamed} column names")

    def reorder_columns(self):
        """Reorder columns according to specified list"""
        if not self.reorder_columns:
            return

        # Normalize column names
        normalized_order = [col.strip().lower().replace(' ', '_') for col in self.reorder_columns]
        available_cols = [c.lower().replace(' ', '_') for c in self.df.columns]

        # Find intersection (preserve order)
        new_order = []
        for col in normalized_order:
            for orig in self.df.columns:
                if col == orig.lower().replace(' ', '_'):
                    new_order.append(orig)
                    break

        # Add missing columns at end
        for orig in self.df.columns:
            if orig not in new_order:
                new_order.append(orig)

        if new_order != list(self.df.columns):
            self.df = self.df[new_order]
            print(f"  Reordered columns: {', '.join(new_order[:5])}...")

    def remove_empty_rows_columns(self):
        """Remove completely empty rows and columns"""
        if not self.remove_empty:
            return

        original_shape = self.df.shape

        # Remove empty rows (all NaN)
        self.df.dropna(how='all', inplace=True)
        rows_removed = original_shape[0] - self.df.shape[0]
        self.stats['rows_removed'] += rows_removed

        # Remove empty columns
        self.df.dropna(axis=1, how='all', inplace=True)
        cols_removed = original_shape[1] - self.df.shape[1]
        self.stats['columns_removed'] += cols_removed

        if rows_removed > 0 or cols_removed > 0:
            print(f"  Removed {rows_removed} empty rows, {cols_removed} empty columns")

    def convert_date_columns(self):
        """Convert specified columns to datetime"""
        if not self.convert_dates:
            return

        for col in self.convert_dates:
            if col not in self.df.columns:
                print(f"  Warning: Column '{col}' not found, skipping")
                continue

            try:
                self.df[col] = pd.to_datetime(self.df[col], errors='coerce')
                self.stats['dates_converted'] += 1
                self.log(f"  Converted '{col}' to datetime")
            except Exception as e:
                print(f"  Warning: Could not convert '{col}': {e}")

    def apply_formats(self, writer):
        """Apply Excel formatting when saving"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        # Save first without formatting
        self.df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()

        # Load workbook to apply formatting
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active

        # Header style
        if self.header_style:
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, col_name in enumerate(self.df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        # Auto column width
        if self.auto_width:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(self.output_file)

    def save(self):
        """Save formatted DataFrame to Excel or CSV"""
        self.log(f"Saving to: {self.output_file}")

        if self.output_file.suffix.lower() == '.csv':
            self.df.to_csv(self.output_file, index=False, encoding='utf-8')
        else:
            # Excel
            if self.auto_width or self.header_style:
                # Need to use ExcelWriter with openpyxl engine for formatting
                with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                    self.apply_formats(writer)
            else:
                self.df.to_excel(self.output_file, index=False, engine='openpyxl')

        print(f"  Saved: {self.df.shape[0]} rows × {self.df.shape[1]} columns")
        self.stats['final_shape'] = self.df.shape

    def print_summary(self):
        """Print formatting summary"""
        print("\n📊 Formatting Summary:")
        print(f"  Original: {self.stats['original_shape'][0]} rows × {self.stats['original_shape'][1]} cols")
        print(f"  Final:    {self.stats['final_shape'][0]} rows × {self.stats['final_shape'][1]} cols")
        print(f"  Renamed:  {self.stats['columns_renamed']} columns")
        print(f"  Removed:  {self.stats['rows_removed']} rows, {self.stats['columns_removed']} columns")
        print(f"  Dates:    {self.stats['dates_converted']} columns converted")

    def run(self):
        """Execute formatting pipeline"""
        self.load_excel()
        self.standardize_column_names()
        self.reorder_columns()
        self.remove_empty_rows_columns()
        self.convert_date_columns()
        self.save()
        self.print_summary()

def main():
    parser = argparse.ArgumentParser(
        description="Excel Formatter - Standardize Excel tables",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Standardize headers (lowercase, underscores)
  %(prog)s data.xlsx --standardize-headers

  # Reorder columns (only specified columns, others at end)
  %(prog)s sales.xlsx --reorder "Date,Product,Quantity,Price"

  # Convert date columns and auto-fit widths
  %(prog)s orders.xlsx --convert-date "Order Date" --convert-date "Ship Date" --auto-width

  # Full formatting: headers, auto-width, remove empty rows
  %(prog)s report.xlsx --standardize-headers --remove-empty --auto-width --header-style

  # Export to CSV instead of Excel
  %(prog)s data.xlsx --standardize-headers --output data_clean.csv

  # Multiple operations
  %(prog)s messy.xlsx --standardize --reorder "Name,Email,Phone" --convert-date "Created" --output clean.xlsx

Notes:
  - Requires pandas + openpyxl: pip install pandas openpyxl
  - Header standardization: removes special chars, replaces spaces with _, lowercases
  - Reorder preserves given column order, appends unspecified columns at end
  - Date conversion uses pandas.to_datetime with errors='coerce' (invalid dates → NaT)
      """
    )
    parser.add_argument(
        "input_file",
        help="Input Excel file (.xlsx or .xls)"
    )
    parser.add_argument(
        "--output",
        help="Output file path (default: input_formatted.xlsx or .csv if extension)"
    )
    parser.add_argument(
        "--standardize-headers",
        action="store_true",
        help="Standardize column names: lowercase, underscores, remove special chars"
    )
    parser.add_argument(
        "--reorder",
        nargs="+",
        help="Column names in desired order (space-separated)"
    )
    parser.add_argument(
        "--remove-empty",
        action="store_true",
        help="Remove completely empty rows and columns"
    )
    parser.add_argument(
        "--convert-date",
        action="append",
        help="Column name to convert to datetime (can use multiple times)"
    )
    parser.add_argument(
        "--date-format",
        help="Custom date format for Excel (e.g., 'YYYY-MM-DD')"
    )
    parser.add_argument(
        "--number-format",
        help="Number format string for Excel (e.g., '#,##0.00')"
    )
    parser.add_argument(
        "--auto-width",
        action="store_true",
        help="Auto-fit column widths in Excel output"
    )
    parser.add_argument(
        "--header-style",
        action="store_true",
        help="Apply styled header row (blue background, white bold text)"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Show detailed processing information"
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"excel_formatter {__version__}"
    )

    args = parser.parse_args()

    formatter = ExcelFormatter(
        input_file=args.input_file,
        output_file=args.output,
        standardize_headers=args.standardize_headers,
        reorder_columns=args.reorder,
        remove_empty=args.remove_empty,
        convert_dates=args.convert_date,
        number_format=args.number_format,
        date_format=args.date_format,
        auto_width=args.auto_width,
        header_style=args.header_style,
        verbose=args.verbose
    )

    try:
        formatter.run()
    except ImportError as e:
        print(f"Error: Missing dependencies. Install with: pip install pandas openpyxl", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
